# trading/excel_reporter.py
# Sends full trade log as Excel to Telegram after every trade

import os
import json
import requests
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

TELEGRAM_TOKEN   = os.environ.get("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
TRADES_FILE      = "data/trades.json"
TRADE_LOG_FILE   = "data/trade_log.json"

# ── Colours ────────────────────────────────────────────────────
NAVY        = "0D1F3C"
BLUE        = "2A6DB5"
GREEN       = "1A6E3E"
GREEN_LIGHT = "E8F5EE"
RED         = "9E2020"
RED_LIGHT   = "FAEAEA"
GOLD        = "B8860B"
GOLD_LIGHT  = "FEF9E7"
GREY        = "F4F6F9"
WHITE       = "FFFFFF"
DARK_TEXT   = "1A1A2E"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="1A1A2E", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def _border():
    thin = Side(style="thin", color="DEE2EC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def build_trade_excel():
    """Build full institutional trade log Excel file."""

    # Load data
    try:
        log = json.load(open(TRADE_LOG_FILE)) if os.path.exists(TRADE_LOG_FILE) else []
    except:
        log = []
    try:
        trades = json.load(open(TRADES_FILE)) if os.path.exists(TRADES_FILE) else []
    except:
        trades = []

    # Merge and deduplicate by date+ticker
    all_trades = log if log else trades

    wb = openpyxl.Workbook()

    # ── Sheet 1: TRADE LOG ─────────────────────────────────────
    ws = wb.active
    ws.title = "Trade Log"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:L1")
    ws["A1"] = "ALPHA TERMINAL — TRADE LOG"
    ws["A1"].font      = _font(bold=True, size=16, color=WHITE)
    ws["A1"].fill      = _fill(NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:L2")
    ws["A2"] = f"Account Z31989293  ·  Fidelity Brokerage  ·  Generated {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}"
    ws["A2"].font      = _font(size=9, color="A0B8D4")
    ws["A2"].fill      = _fill(NAVY)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8

    # Headers
    headers = [
        "Date", "Action", "Ticker", "Sector", "Price",
        "Shares", "Cost / P&L $", "Return %",
        "Macro Regime", "Probability Score",
        "Why We Acted", "Lesson Learned"
    ]
    col_widths = [14, 8, 8, 16, 10, 8, 14, 10, 18, 10, 45, 45]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = _font(bold=True, size=9, color=WHITE)
        cell.fill      = _fill(BLUE)
        cell.alignment = _center()
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 28

    # Data rows
    wins = losses = 0
    total_pnl = 0

    for i, t in enumerate(all_trades[:200]):
        row  = i + 5
        action  = t.get("action") or t.get("type", "")
        is_buy  = action == "BUY"
        pnl     = t.get("pnl")
        pnl_pct = t.get("pnl_pct")
        cost    = (t.get("price", 0) or 0) * (t.get("shares", 0) or 0)

        if not is_buy:
            if pnl and pnl >= 0:
                wins += 1
                total_pnl += pnl or 0
            elif pnl and pnl < 0:
                losses += 1
                total_pnl += pnl or 0

        # Row background
        if is_buy:
            row_fill = _fill("EEF4FC")
        elif pnl and pnl >= 0:
            row_fill = _fill(GREEN_LIGHT)
        elif pnl and pnl < 0:
            row_fill = _fill(RED_LIGHT)
        else:
            row_fill = _fill(GREY if i % 2 == 0 else WHITE)

        values = [
            (t.get("date", "")[:10],                             _center()),
            (action,                                              _center()),
            (t.get("ticker", ""),                                 _center()),
            (t.get("sector", ""),                                 _center()),
            (t.get("price"),                                      _center()),
            (t.get("shares"),                                     _center()),
            (cost if is_buy else pnl,                             _center()),
            (pnl_pct,                                             _center()),
            (t.get("macro_env") or t.get("macro_alignment", ""), _left()),
            (t.get("probability_score"),                          _center()),
            (t.get("reasoning", "")[:300],                        _left()),
            (t.get("sell_reason") or t.get("lesson", ""),         _left()),
        ]

        for col, (val, align) in enumerate(values, 1):
            cell           = ws.cell(row=row, column=col, value=val)
            cell.fill      = row_fill
            cell.border    = _border()
            cell.alignment = align

            # Colour specific cells
            if col == 2:  # Action
                cell.font = _font(bold=True, size=9,
                    color=GREEN if is_buy else (GREEN if (pnl or 0) >= 0 else RED))
            elif col == 3:  # Ticker
                cell.font = _font(bold=True, size=10, color=NAVY)
            elif col == 7:  # P&L $
                cell.font = _font(bold=True, size=9,
                    color=GREEN if (pnl or 0) >= 0 else RED)
                if not is_buy and val is not None:
                    cell.number_format = '+#,##0.00;-#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
            elif col == 8:  # Return %
                cell.font = _font(bold=True, size=9,
                    color=GREEN if (pnl_pct or 0) >= 0 else RED)
                if val is not None:
                    cell.number_format = '+0.00%;-0.00%'
            elif col == 5:  # Price
                cell.number_format = '$#,##0.00'
            else:
                cell.font = _font(size=9)

        ws.row_dimensions[row].height = 32

    # ── Sheet 2: SUMMARY ──────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "PERFORMANCE SUMMARY"
    ws2["A1"].font      = _font(bold=True, size=14, color=WHITE)
    ws2["A1"].fill      = _fill(NAVY)
    ws2["A1"].alignment = _center()
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells("A2:D2")
    ws2["A2"] = f"Generated {datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}"
    ws2["A2"].font      = _font(size=9, color="A0B8D4")
    ws2["A2"].fill      = _fill(NAVY)
    ws2["A2"].alignment = _center()

    metrics = [
        ("", ""),
        ("METRIC",                "VALUE"),
        ("Total Trades",          len([t for t in all_trades if (t.get("action") or t.get("type","")) == "SELL"])),
        ("Winning Trades",        wins),
        ("Losing Trades",         losses),
        ("Win Rate",              f"{wins/(wins+losses)*100:.1f}%" if wins+losses > 0 else "N/A"),
        ("Total Realised P&L",    f"${total_pnl:+,.2f}"),
        ("Starting Capital",      "$183,000"),
        ("",                      ""),
        ("Best Sector",           _best_sector(all_trades)),
        ("Worst Sector",          _worst_sector(all_trades)),
    ]

    for i, (label, value) in enumerate(metrics, 3):
        ws2.row_dimensions[i].height = 24
        for col, val in enumerate([label, value], 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = _border()
            ws2.column_dimensions[get_column_letter(col)].width = 25
            if i == 4:  # header row
                cell.font = _font(bold=True, size=10, color=WHITE)
                cell.fill = _fill(BLUE)
                cell.alignment = _center()
            elif label and value:
                cell.fill = _fill(GREY if i % 2 == 0 else WHITE)
                cell.font = _font(bold=(col == 1), size=10,
                    color=GREEN if ("+" in str(value) and col == 2)
                          else RED if ("-" in str(value) and col == 2 and "Worst" not in label)
                          else DARK_TEXT)
                cell.alignment = _center() if col == 2 else _left()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def _best_sector(trades):
    sector_pnl = {}
    for t in trades:
        s   = t.get("sector", "OTHER") or "OTHER"
        pnl = t.get("pnl", 0) or 0
        sector_pnl[s] = sector_pnl.get(s, 0) + pnl
    if not sector_pnl:
        return "N/A"
    return max(sector_pnl, key=sector_pnl.get)

def _worst_sector(trades):
    sector_pnl = {}
    for t in trades:
        s   = t.get("sector", "OTHER") or "OTHER"
        pnl = t.get("pnl", 0) or 0
        sector_pnl[s] = sector_pnl.get(s, 0) + pnl
    if not sector_pnl:
        return "N/A"
    return min(sector_pnl, key=sector_pnl.get)

def send_trade_excel(trigger_action="", trigger_ticker=""):
    """Build Excel and send to Telegram."""
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("[Excel] No Telegram credentials — skipping")
        return
    try:
        excel_bytes = build_trade_excel()
        now         = datetime.utcnow().strftime("%d%b%Y")
        filename    = f"ALPHA_TradeLog_{now}.xlsx"
        caption     = (
            f"ALPHA TRADE LOG — Updated\n"
            f"Triggered by: {trigger_action} ${trigger_ticker}\n"
            f"{datetime.utcnow().strftime('%d %B %Y %H:%M UTC')}"
        )
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
            files={"document": (filename, excel_bytes,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=30,
        )
        print(f"[Excel] Sent trade log Excel to Telegram")
    except Exception as e:
        print(f"[Excel] Error: {e}")
