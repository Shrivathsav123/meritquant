#!/usr/bin/env python3
"""
MeritQuant — Autonomous AI Trader
Claude Opus 4.8 decision engine with macro brain and trade memory.
Trade windows: 9:35 AM ET (open) and 3:30 PM ET (pre-close).
GitHub Actions TRADE_MODE env var prevents timing drift failures.
"""

import os, json, re, time, logging, io, requests
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

import anthropic
import yfinance as yf
from trading import email_alerts
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    datefmt="%Y-%m-%d %H:%M:%S")
log = logging.getLogger("ai_trader")

# ── Constants ─────────────────────────────────────────────────────────────────
PORTFOLIO_SIZE     = 183_000.0
MAX_POSITION_PCT   = 0.10
MAX_POSITION_USD   = 18_000          # flat per-position size — no conviction-based tiers
MAX_OPEN_POSITIONS = 10
STOP_LOSS_PCT      = 0.08
TAKE_PROFIT_PCT    = 0.20
MIN_CONVICTION     = 5      # Don't trade below this score

SIGNALS_FILE   = "data/signals.json"
POSITIONS_FILE = "data/positions.json"
TRADE_LOG_FILE = "data/trade_log.json"
MEMORY_FILE    = "data/memory.json"
REPORTS_DIR    = "reports"
TELEGRAM_BASE  = "https://api.telegram.org"

# ── Environment ───────────────────────────────────────────────────────────────
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
TELEGRAM_TOKEN    = os.environ["TELEGRAM_TOKEN"]
TELEGRAM_CHAT_ID  = os.environ["TELEGRAM_CHAT_ID"]
FRED_API_KEY      = os.environ.get("FRED_API_KEY", "")
# TRADE_MODE is set by GitHub Actions based on which cron triggered.
# "true"  → dedicated trade window schedule fired → always trade
# "false" → scan-only schedule fired → skip AI trader
TRADE_MODE = os.environ.get("TRADE_MODE", "false").lower() == "true"


def get_scan_context():
    import os as _os
    from datetime import datetime as _dt, timezone as _tz
    path = 'data/alpha_terminal_scan.json'
    try:
        if not _os.path.exists(path):
            return ""
        with open(path, 'r') as f:
            scan = json.load(f)
        scan_time = _dt.fromisoformat(scan.get('scan_time', '').replace('Z', '+00:00'))
        age_hours = (_dt.now(_tz.utc) - scan_time).total_seconds() / 3600
        if age_hours > 2:
            return f"[SCAN DATA STALE - {age_hours:.1f}hrs old, proceed with caution]"
        return f"""
=== MERITQUANT TECHNICAL SCAN ({scan['scan_type']} | {scan['scan_time']}) ===

MACRO REGIME: {scan['macro']['regime']}

VIX: {scan['macro']['vix']} | DXY: {scan['macro']['dxy']} | \nYield Spread (10Y-2Y): {scan['macro']['spread_10y_2y_bps']}bps

HYG: {scan['macro']['hyg']} | LQD: {scan['macro']['lqd']}

Position Size Cap: {scan['macro']['position_size_cap_pct']}% \n(${scan['macro']['position_size_cap_usd']:,.0f} max per trade)

Macro Notes: {scan['macro']['macro_notes']}

SECTOR ROTATION:

Leading (BUY SIDE): {scan['sector_rotation']['leading_sector']} \n  ({scan['sector_rotation']['leading_avg_pct']:+.2f}%)

Lagging (AVOID): {scan['sector_rotation']['lagging_sector']} \n  ({scan['sector_rotation']['lagging_avg_pct']:+.2f}%)

DXY Signal: {scan['sector_rotation']['dxy_equity_signal']}

Thesis: {scan['sector_rotation']['rotation_thesis']}

ACTIVE FVG SETUPS (ranked by conviction):

{chr(10).join([
    f"  #{i+1} {s['symbol']} {s['direction']} | "
    f"Entry: ${s['entry_price']} | Stop: ${s['stop_price']} | "
    f"Target: ${s['target_1']} | R:R {s['risk_reward']:.1f}x | "
    f"Gates: {s['gates_cleared']}/9 | Size: {s['position_size_pct']}% "
    f"(${s['position_size_usd']:,.0f}) | {s['thesis']}"
    for i, s in enumerate(scan.get('setups', []))
])}

WATCHLIST (not ready yet):

{chr(10).join([
    f"  {w['symbol']} - Alert at ${w['alert_price']} | Missing: {w['missing']}"
    for w in scan.get('no_trade_watchlist', [])
])}

SCANNER DIRECTIVE FOR THIS DECISION:

{scan['ai_directive']}

=== END SCAN DATA ===

"""
    except Exception as e:
        return f"[Scan data unavailable: {e}]"


# ─────────────────────────────────────────────────────────────────────────────
# 1. UTILITY HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def load_json(path: str, default=None):
    p = Path(path)
    if p.exists():
        try:
            return json.loads(p.read_text())
        except Exception:
            pass
    return default if default is not None else {}

def save_json(path: str, data):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(data, indent=2, default=str))

PORTFOLIO_JSON_FILE = "data/portfolio.json"

def sync_portfolio_json(portfolio: dict):
    """Write portfolio.json in the format the meritquant-app frontend expects."""
    positions_dict = {}
    for p in portfolio.get("positions", []):
        ticker = p["ticker"]
        cost   = p.get("cost_basis_total", 0)
        value  = p.get("current_value", cost)
        pnl    = p.get("unrealised_pnl", 0)
        positions_dict[ticker] = {
            "ticker":      ticker,
            "name":        p.get("name", ticker),
            "shares":      p.get("shares", 0),
            "entry_price": p.get("entry_price", 0),
            "current_price": p.get("current_price", p.get("entry_price", 0)),
            "cost":        round(cost, 2),
            "value":       round(value, 2),
            "pnl":         round(pnl, 2),
            "pnl_pct":     round(p.get("unrealised_pct", 0) * 100, 2),
            "stop_loss":   round(p.get("entry_price", 0) * (1 - STOP_LOSS_PCT), 2),
            "trade_type":  "swing",
            "score":       p.get("conviction", 0),
            "reasoning":   p.get("thesis_summary", ""),
            "entry_date":  p.get("entry_date", ""),
        }

    total_value = portfolio.get("total_value", PORTFOLIO_SIZE)
    cash        = portfolio.get("cash", PORTFOLIO_SIZE)
    pnl_total   = total_value - PORTFOLIO_SIZE
    app_data = {
        "balance":      PORTFOLIO_SIZE,
        "cash":         round(cash, 2),
        "positions":    positions_dict,
        "total_value":  round(total_value, 2),
        "pnl":          round(pnl_total, 2),
        "pnl_pct":      round(pnl_total / PORTFOLIO_SIZE * 100, 2),
        "updated":      datetime.utcnow().isoformat(),
        "trades_count": len(portfolio.get("positions", [])),
    }
    save_json(PORTFOLIO_JSON_FILE, app_data)

def send_telegram(text: str, parse_mode: str = "HTML") -> bool:
    url = f"{TELEGRAM_BASE}/bot{TELEGRAM_TOKEN}/sendMessage"
    try:
        r = requests.post(url, json={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": text,
            "parse_mode": parse_mode
        }, timeout=15)
        return r.ok
    except Exception as e:
        log.error(f"Telegram text failed: {e}")
        return False

def send_telegram_doc(file_bytes: bytes, filename: str, caption: str = "") -> bool:
    url = f"{TELEGRAM_BASE}/bot{TELEGRAM_TOKEN}/sendDocument"
    try:
        r = requests.post(url,
            data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption},
            files={"document": (filename, file_bytes)},
            timeout=30)
        return r.ok
    except Exception as e:
        log.error(f"Telegram doc failed: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# 2. TRADE WINDOW CHECK  (fix: TRADE_MODE env var + ±15 min buffer)
# ─────────────────────────────────────────────────────────────────────────────
def is_trade_window() -> bool:
    """
    Returns True when AI trader should execute decisions.

    Priority 1: TRADE_MODE env var (set by GitHub Actions workflow).
      - Dedicated trade-window crons set TRADE_MODE=true
      - Scan-only crons set TRADE_MODE=false
      This eliminates GitHub Actions timing drift as a failure mode.

    Priority 2: Time-based fallback with ±15 min buffer around
      9:35 AM ET (open) and 3:30 PM ET (pre-close).
      Handles manual workflow_dispatch runs and local execution.
    """
    # Hard override from workflow scheduler
    if TRADE_MODE:
        log.info("TRADE_MODE=true — proceeding with AI trade decisions.")
        return True

    # Time-based fallback
    ET = timezone(timedelta(hours=-4))   # EDT (UTC-4); change to -5 Nov–Mar
    now_et = datetime.now(ET)

    if now_et.weekday() >= 5:            # Saturday=5, Sunday=6
        log.info("Weekend — markets closed. Skipping.")
        return False

    now_min   = now_et.hour * 60 + now_et.minute
    OPEN_WIN  = 9  * 60 + 35            # 575 minutes
    CLOSE_WIN = 15 * 60 + 30            # 930 minutes
    BUFFER    = 15                       # ±15 min handles GH Actions drift

    in_open  = abs(now_min - OPEN_WIN)  <= BUFFER
    in_close = abs(now_min - CLOSE_WIN) <= BUFFER

    if in_open or in_close:
        label = "OPEN" if in_open else "PRE-CLOSE"
        log.info(f"Trade window: {label} at {now_et.strftime('%H:%M')} ET")
        return True

    log.info(f"Scan-only window at {now_et.strftime('%H:%M')} ET — skipping AI trades.")
    return False


# ─────────────────────────────────────────────────────────────────────────────
# 3. MACRO BRAIN
# ─────────────────────────────────────────────────────────────────────────────
def fetch_fred(series_id: str) -> Optional[float]:
    """Fetch latest value from FRED public CSV endpoint."""
    try:
        url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
        if FRED_API_KEY:
            url += f"&api_key={FRED_API_KEY}"
        r = requests.get(url, timeout=12)
        for line in reversed(r.text.strip().split("\n")[1:]):
            parts = line.split(",")
            if len(parts) == 2 and parts[1].strip() not in ("", "."):
                return float(parts[1].strip())
    except Exception as e:
        log.warning(f"FRED {series_id}: {e}")
    return None

def fetch_price(ticker: str) -> Optional[float]:
    """Fetch latest close from Yahoo Finance."""
    try:
        hist = yf.Ticker(ticker).history(period="2d")
        if not hist.empty:
            return float(hist["Close"].iloc[-1])
    except Exception as e:
        log.warning(f"Yahoo {ticker}: {e}")
    return None

def build_macro_context() -> dict:
    log.info("Building macro brain context...")

    vix       = fetch_price("^VIX")
    dxy       = fetch_price("DX-Y.NYB")
    spy       = fetch_price("SPY")
    t10y2y    = fetch_fred("T10Y2Y")        # Yield curve
    hy_spread = fetch_fred("BAMLH0A0HYM2") # HY credit OAS bps
    t10y      = fetch_fred("DGS10")         # 10Y yield
    t2y       = fetch_fred("DGS2Y")         # 2Y yield
    fed_rate  = fetch_fred("FEDFUNDS")      # Fed funds

    # ── Regime classification ─────────────────────────────────────────────────
    regime  = "NEUTRAL"
    reasons = []

    if vix is not None:
        if vix > 30:
            regime = "RISK-OFF"
            reasons.append(f"VIX elevated at {vix:.1f} — fear spike")
        elif vix < 18:
            reasons.append(f"VIX calm at {vix:.1f} — supportive for longs")
        else:
            reasons.append(f"VIX neutral at {vix:.1f}")

    if t10y2y is not None:
        if t10y2y < 0:
            if regime != "RISK-OFF":
                regime = "CAUTION"
            reasons.append(f"Yield curve inverted at {t10y2y:.2f}% — recession signal")
        else:
            reasons.append(f"Yield curve positive at +{t10y2y:.2f}%")

    if hy_spread is not None:
        if hy_spread > 500:
            regime = "RISK-OFF"
            reasons.append(f"HY spreads wide at {hy_spread:.0f}bps — credit stress")
        elif hy_spread < 300:
            reasons.append(f"HY spreads tight at {hy_spread:.0f}bps — credit healthy")
        else:
            reasons.append(f"HY spreads neutral at {hy_spread:.0f}bps")

    if regime == "NEUTRAL" and vix and vix < 20 and (t10y2y or 0) > 0:
        regime = "RISK-ON"

    macro = {
        "timestamp": datetime.utcnow().isoformat(),
        "regime": regime,
        "regime_reasons": reasons,
        "indicators": {
            "VIX":       {"value": vix,       "label": "CBOE Volatility Index"},
            "DXY":       {"value": dxy,       "label": "US Dollar Index"},
            "SPY":       {"value": spy,       "label": "S&P 500 ETF"},
            "T10Y2Y":    {"value": t10y2y,    "label": "Yield Curve 10Y–2Y (%)"},
            "HY_SPREAD": {"value": hy_spread, "label": "HY Credit OAS (bps)"},
            "T10Y":      {"value": t10y,      "label": "10Y Treasury (%)"},
            "T2Y":       {"value": t2y,       "label": "2Y Treasury (%)"},
            "FED_RATE":  {"value": fed_rate,  "label": "Fed Funds Rate (%)"},
        }
    }

    log.info(f"Macro: {regime} | VIX={vix} | Curve={t10y2y} | HY={hy_spread}bps")
    return macro


# ─────────────────────────────────────────────────────────────────────────────
# 4. BALANCE SHEET ANALYSIS (SEC EDGAR XBRL)
# ─────────────────────────────────────────────────────────────────────────────
_CIK_MAP: dict = {}


def _load_cik_map() -> dict:
    global _CIK_MAP
    if _CIK_MAP:
        return _CIK_MAP
    try:
        r = requests.get(
            "https://www.sec.gov/files/company_tickers.json",
            headers={"User-Agent": "MeritQuant research@meritquant.com"},
            timeout=15,
        )
        r.raise_for_status()
        _CIK_MAP = {v["ticker"].upper(): str(v["cik_str"]) for v in r.json().values()}
    except Exception as e:
        log.warning(f"SEC CIK map: {e}")
    return _CIK_MAP


def _annual_trend(facts_json: dict, concept: str, n: int = 3) -> list:
    """Return last n 10-K values for a US-GAAP concept, most recent first."""
    try:
        vals = (facts_json.get("facts", {})
                          .get("us-gaap", {})
                          .get(concept, {})
                          .get("units", {})
                          .get("USD", []))
        annual = [v for v in vals
                  if v.get("form") in ("10-K", "10-K/A") and v.get("val") is not None]
        seen, out = set(), []
        for v in sorted(annual, key=lambda x: x.get("end", ""), reverse=True):
            yr = v["end"][:4]
            if yr not in seen:
                seen.add(yr)
                out.append(v["val"])
            if len(out) == n:
                break
        return out
    except Exception:
        return []


def fetch_balance_sheet(ticker: str) -> dict:
    """3-tier fallback: SEC EDGAR → yfinance balance_sheet → yfinance info.
    Always returns a populated score (0-10) and summary string."""

    def _compute(cash, total_debt, equity, de_ratio, retained, treasury, preferred):
        score = 0
        if cash:
            score += 2 if cash >= 5_000_000_000 else 1 if cash >= 500_000_000 else 0
        if de_ratio is not None:
            score += 2 if de_ratio < 0.3 else 1 if de_ratio < 0.7 else 0
        elif total_debt == 0 and equity and equity > 0:
            score += 2
        if retained and retained[0] is not None and retained[0] > 0:
            score += 1
        if len(retained) >= 2 and None not in retained[:2]:
            if retained[0] > retained[1]:
                score += 1
        if len(retained) == 3 and None not in retained:
            if retained[1] > retained[2]:
                score += 1
        score += 1 if (preferred is None or preferred == 0) else 0
        score += 1 if (treasury and treasury > 0) else 0
        score += 1 if (equity and equity > 0) else 0
        score = min(score, 10)

        cash_s = f"${cash/1e9:.2f}B" if cash else "N/A"
        debt_s = f"${total_debt/1e9:.2f}B" if total_debt else "$0B"
        de_s   = f"{de_ratio:.2f}" if de_ratio is not None else "N/A"
        if len(retained) >= 2 and None not in retained[:2]:
            pct  = (retained[0] - retained[1]) / abs(retained[1]) * 100 if retained[1] else 0
            re_s = f"growing (+{pct:.0f}%)" if retained[0] > retained[1] else f"declining ({pct:.0f}%)"
        elif retained and retained[0] is not None:
            re_s = f"${retained[0]/1e9:.1f}B"
        else:
            re_s = "N/A"
        summary = (f"Cash {cash_s} vs Debt {debt_s} — D/E ratio {de_s} — "
                   f"Retained earnings {re_s} — Buffett score {score}/10")
        return score, summary

    # Tier 1: SEC EDGAR
    cik = _load_cik_map().get(ticker.upper())
    if cik:
        try:
            url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{int(cik):010d}.json"
            r = requests.get(url, headers={"User-Agent": "MeritQuant research@meritquant.com"}, timeout=20)
            if r.status_code == 200:
                facts = r.json()
                def one(concept):
                    vals = _annual_trend(facts, concept, 1)
                    return vals[0] if vals else None
                cash      = one("CashAndCashEquivalentsAtCarryingValue")
                lt_debt   = one("LongTermDebt") or 0
                st_debt   = one("ShortTermBorrowings") or one("ShortTermDebt") or 0
                equity    = one("StockholdersEquity")
                retained  = _annual_trend(facts, "RetainedEarningsAccumulatedDeficit", 3)
                treasury  = one("TreasuryStockValue")
                preferred = one("PreferredStockValue")
                total_debt = (lt_debt or 0) + (st_debt or 0)
                de_ratio   = (total_debt / equity) if (equity and equity > 0) else None
                if cash or equity:
                    score, summary = _compute(cash, total_debt, equity, de_ratio, retained, treasury, preferred)
                    return {"ticker": ticker, "source": "edgar", "cash": cash,
                            "total_debt": total_debt, "lt_debt": lt_debt, "st_debt": st_debt,
                            "equity": equity,
                            "de_ratio": round(de_ratio, 3) if de_ratio is not None else None,
                            "retained": retained, "buybacks": bool(treasury and treasury > 0),
                            "preferred": bool(preferred and preferred > 0),
                            "score": score, "summary": summary}
        except Exception as e:
            log.warning(f"SEC EDGAR {ticker}: {e}")

    # Tier 2: yfinance balance_sheet DataFrame
    try:
        yt = yf.Ticker(ticker)
        bs = yt.balance_sheet
        if bs is not None and not bs.empty:
            def _row(df, *keys):
                for k in keys:
                    if k in df.index:
                        vals = df.loc[k].dropna()
                        return float(vals.iloc[0]) if len(vals) else None
                return None
            def _trend(df, *keys):
                for k in keys:
                    if k in df.index:
                        vals = df.loc[k].dropna()
                        return [float(v) for v in vals.iloc[:3]]
                return []
            cash     = _row(bs, "Cash And Cash Equivalents",
                            "Cash Cash Equivalents And Short Term Investments",
                            "Cash And Short Term Investments")
            debt     = _row(bs, "Total Debt", "Long Term Debt And Capital Lease Obligation", "Long Term Debt")
            equity   = _row(bs, "Stockholders Equity", "Common Stockholders Equity",
                            "Total Equity Gross Minority Interest")
            retained = _trend(bs, "Retained Earnings", "Retained Earnings Deficit", "Accumulated Deficit")
            total_debt = debt or 0
            de_ratio   = (total_debt / equity) if (equity and equity > 0) else None
            if cash or equity:
                score, summary = _compute(cash, total_debt, equity, de_ratio, retained, None, None)
                return {"ticker": ticker, "source": "yfinance_bs", "cash": cash,
                        "total_debt": total_debt, "equity": equity,
                        "de_ratio": round(de_ratio, 3) if de_ratio is not None else None,
                        "retained": retained, "buybacks": False, "preferred": False,
                        "score": score, "summary": summary}
    except Exception as e:
        log.warning(f"yfinance balance_sheet {ticker}: {e}")

    # Tier 3: yfinance .info
    try:
        info = yf.Ticker(ticker).info
        cash      = info.get("totalCash")
        total_debt = info.get("totalDebt") or 0
        de_raw    = info.get("debtToEquity")  # yfinance: D/E × 100
        de_ratio  = de_raw / 100 if de_raw is not None else None
        equity    = (total_debt / de_ratio) if (de_ratio and de_ratio > 0) else None
        score, summary = _compute(cash, total_debt, equity, de_ratio, [], None, None)
        return {"ticker": ticker, "source": "yfinance_info", "cash": cash,
                "total_debt": total_debt, "equity": equity,
                "de_ratio": round(de_ratio, 3) if de_ratio is not None else None,
                "retained": [], "buybacks": False, "preferred": False,
                "score": score, "summary": summary}
    except Exception as e:
        log.warning(f"yfinance info {ticker}: {e}")

    # All sources failed — return zero-data record with generated summary
    score, summary = _compute(None, 0, None, None, [], None, None)
    return {"ticker": ticker, "source": "none", "cash": None, "total_debt": 0,
            "equity": None, "de_ratio": None, "retained": [], "buybacks": False,
            "preferred": False, "score": score, "summary": summary}


# ─────────────────────────────────────────────────────────────────────────────
# 5. TRADE MEMORY SYSTEM
# ─────────────────────────────────────────────────────────────────────────────
def load_memory() -> dict:
    return load_json(MEMORY_FILE, {
        "losing_trades": [],
        "winning_patterns": [],
        "macro_lessons": [],
        "last_updated": None
    })

def save_memory(memory: dict):
    memory["last_updated"] = datetime.utcnow().isoformat()
    save_json(MEMORY_FILE, memory)

def update_memory(trade: dict, memory: dict) -> dict:
    pnl = trade.get("pnl_pct", 0)
    if pnl < -0.05:
        memory["losing_trades"].append({
            "ticker":     trade.get("ticker"),
            "date":       trade.get("exit_date", "")[:10],
            "pnl_pct":    round(pnl, 4),
            "why_failed": trade.get("exit_reason", ""),
            "thesis":     trade.get("thesis_summary", ""),
            "lesson":     trade.get("lesson", "Verify thesis confirmation before full sizing.")
        })
        memory["losing_trades"] = memory["losing_trades"][-20:]
    if pnl > 0.08:
        memory["winning_patterns"].append({
            "ticker":          trade.get("ticker"),
            "date":            trade.get("exit_date", "")[:10],
            "pnl_pct":         round(pnl, 4),
            "thesis_summary":  trade.get("thesis_summary", ""),
            "catalyst":        trade.get("catalyst", ""),
            "technical_setup": trade.get("technical_setup", ""),
            "conviction":      trade.get("conviction", ""),
            "entry_date":      trade.get("entry_date", "")[:10],
        })
        memory["winning_patterns"] = memory["winning_patterns"][-20:]
    return memory

def memory_to_prompt(memory: dict) -> str:
    parts = []
    if memory.get("losing_trades"):
        parts.append("PAST LOSING TRADES — DO NOT REPEAT:")
        for t in memory["losing_trades"][-5:]:
            parts.append(f"  • {t['ticker']} ({t['date']}) | {t['pnl_pct']*100:.1f}% | "
                         f"Why: {t['why_failed']} | Lesson: {t['lesson']}")
    if memory.get("winning_patterns"):
        parts.append("\nWINNING SETUPS TO REPLICATE:")
        for p in memory["winning_patterns"][-5:]:
            parts.append(
                f"  • {p['ticker']} ({p.get('date','')}) | +{p['pnl_pct']*100:.1f}% | "
                f"Catalyst: {p.get('catalyst','')} | "
                f"Setup: {p.get('technical_setup', p.get('setup',''))} | "
                f"Conviction: {p.get('conviction','—')}/10"
            )
        parts.append(
            "\nREPLICATE THESE WINNING SETUPS — when you see a current signal matching a past winner "
            "in sector, RSI range, catalyst tier, and macro regime, increase your conviction by 1 point "
            "and prioritise entry."
        )
    return "\n".join(parts) if parts else "No significant trade history recorded yet."


# ─────────────────────────────────────────────────────────────────────────────
# 5. PORTFOLIO STATE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────
def load_positions() -> dict:
    if Path(POSITIONS_FILE).exists():
        return load_json(POSITIONS_FILE, {})
    # Fresh start — try to carry forward cash from the app-facing portfolio.json
    # so accumulated P&L is never lost on a first run.
    legacy = load_json(PORTFOLIO_JSON_FILE, {})
    starting_cash = legacy.get("cash", legacy.get("total_value", PORTFOLIO_SIZE))
    return {
        "positions": [],
        "cash": float(starting_cash),
        "total_value": float(starting_cash),
        "last_updated": None,
    }

def load_trade_log() -> list:
    return load_json(TRADE_LOG_FILE, [])

def update_prices(portfolio: dict) -> dict:
    total = portfolio.get("cash", 0)
    for pos in portfolio.get("positions", []):
        price = fetch_price(pos["ticker"])
        if price:
            pos["current_price"]  = price
            pos["current_value"]  = price * pos["shares"]
            pos["unrealised_pnl"] = pos["current_value"] - pos["cost_basis_total"]
            pos["unrealised_pct"] = pos["unrealised_pnl"] / pos["cost_basis_total"]
        total += pos.get("current_value", pos["cost_basis_total"])
    portfolio["total_value"]  = total
    portfolio["last_updated"] = datetime.utcnow().isoformat()
    return portfolio

def close_pos(pos: dict, reason: str) -> dict:
    return {
        "type":             "CLOSE",
        "ticker":           pos["ticker"],
        "shares":           pos["shares"],
        "entry_price":      pos["entry_price"],
        "exit_price":       pos.get("current_price", pos["entry_price"]),
        "cost_basis_total": pos["cost_basis_total"],
        "exit_value":       pos.get("current_value", pos["cost_basis_total"]),
        "pnl_usd":          pos.get("unrealised_pnl", 0),
        "pnl_pct":          pos.get("unrealised_pct", 0),
        "exit_reason":      reason,
        "thesis_summary":   pos.get("thesis_summary", ""),
        "catalyst":         pos.get("catalyst", ""),
        "technical_setup":  pos.get("technical_setup", ""),
        "entry_date":       pos.get("entry_date", ""),
        "exit_date":        datetime.utcnow().isoformat(),
        "trade_date":       datetime.utcnow().isoformat(),
    }

def run_sl_tp(portfolio: dict, trade_log: list, memory: dict):
    """Auto-exit positions at stop loss (−8%) or take profit (+20%)."""
    exits, remaining = [], []
    for pos in portfolio.get("positions", []):
        pnl = pos.get("unrealised_pct", 0)
        if pnl <= -STOP_LOSS_PCT:
            reason = f"STOP LOSS at {pnl*100:.1f}%"
            t = close_pos(pos, reason)
            trade_log.append(t)
            memory = update_memory(t, memory)
            portfolio["cash"] += pos.get("current_value", pos["cost_basis_total"])
            exits.append((pos["ticker"], reason, pnl))
            log.warning(f"SL HIT: {pos['ticker']} {pnl*100:.1f}%")
        elif pnl >= TAKE_PROFIT_PCT:
            reason = f"TAKE PROFIT at +{pnl*100:.1f}%"
            t = close_pos(pos, reason)
            trade_log.append(t)
            portfolio["cash"] += pos.get("current_value", pos["cost_basis_total"])
            exits.append((pos["ticker"], reason, pnl))
            log.info(f"TP HIT: {pos['ticker']} +{pnl*100:.1f}%")
        else:
            remaining.append(pos)
    portfolio["positions"] = remaining
    return portfolio, trade_log, memory, exits


# ─────────────────────────────────────────────────────────────────────────────
# 6. CLAUDE OPUS 4.8 DECISION ENGINE
# ─────────────────────────────────────────────────────────────────────────────
def build_prompt(signals: list, macro: dict, portfolio: dict, memory: dict) -> str:
    ind = macro.get("indicators", {})

    def v(key):
        val = ind.get(key, {}).get("value")
        return f"{val:.2f}" if val is not None else "N/A"

    # ── Macro block ────────────────────────────────────────────────────────
    macro_block = (
        f"MACRO ENVIRONMENT — {macro['regime']} REGIME\n"
        + "\n".join(f"  ▸ {r}" for r in macro.get("regime_reasons", []))
        + f"\n  VIX: {v('VIX')}  |  DXY: {v('DXY')}  |  SPY: ${v('SPY')}"
        f"\n  Yield Curve (10Y-2Y): {v('T10Y2Y')}%  |  HY Credit Spread: {v('HY_SPREAD')}bps"
        f"\n  10Y Treasury: {v('T10Y')}%  |  2Y Treasury: {v('T2Y')}%  |  Fed Funds: {v('FED_RATE')}%"
    )

    # ── Portfolio block (with live SL/TP levels) ───────────────────────────
    positions = portfolio.get("positions", [])
    port_block = (
        f"Cash: ${portfolio.get('cash', 0):,.0f}  |  "
        f"Total: ${portfolio.get('total_value', 0):,.0f}  |  "
        f"Positions: {len(positions)}/{MAX_OPEN_POSITIONS}\n"
    )
    for p in positions:
        entry = p.get("entry_price", 0)
        curr  = p.get("current_price", entry)
        sl    = entry * (1 - STOP_LOSS_PCT)
        tp    = entry * (1 + TAKE_PROFIT_PCT)
        port_block += (
            f"  {p['ticker']}: {p['shares']:.1f}sh @ ${entry:.2f} → ${curr:.2f} | "
            f"SL ${sl:.2f} | TP ${tp:.2f} | "
            f"P&L {p.get('unrealised_pct', 0)*100:+.1f}% | "
            f"Thesis: {p.get('thesis_summary', '')[:60]}\n"
        )

    # ── Signal block (scanner score + balance sheet + technical) ──────────
    top_signals = sorted(signals, key=lambda x: x.get("score", 0), reverse=True)[:12]
    hdr = (f"  {'TICKER':<6} | {'Scan':>5} | {'RSI':>5} | {'BS':>5} | "
           f"{'D/E':>5} | {'Cash$B':>7} | {'Sig':<4} | {'Sector':<16} | Headline")
    sep = "  " + "-" * 110
    rows = [hdr, sep]
    for s in top_signals:
        bs_score = s.get("bs_score")
        bs_str   = f"{bs_score}/10" if bs_score is not None else " N/A"
        de_val   = s.get("bs_de")
        de_str   = f"{de_val:.2f}" if de_val is not None else "  N/A"
        cash_b   = s.get("bs_cash")
        cash_str = f"${cash_b:.1f}B" if cash_b is not None else "    N/A"
        try:
            rsi_str = f"{float(s.get('rsi', 'N/A')):.1f}"
        except (TypeError, ValueError):
            rsi_str = "  N/A"
        rows.append(
            f"  {s.get('ticker','?'):<6} | {s.get('score',0):5.1f} | {rsi_str:>5} | "
            f"{bs_str:>5} | {de_str:>5} | {cash_str:>7} | "
            f"{s.get('signal','?'):<4} | {s.get('sector','?'):<16} | "
            f"{s.get('news_headline','')[:65]}"
        )
    sig_block = "TOP SIGNALS (Scanner → Balance Sheet → Technical):\n" + "\n".join(rows)

    window = "PRE-CLOSE (3:30 PM ET)" if datetime.utcnow().hour >= 19 else "OPEN (9:35 AM ET)"

    return get_scan_context() + "\n\n" + f"""You are the Head of Quantitative Equity Strategy at MeritQuant — managing a ${PORTFOLIO_SIZE:,.0f} institutional paper portfolio modelled on an NGO endowment trust. Trade window: {window}.

INVESTMENT MANDATE:
- Buffett balance-sheet discipline: prefer cash-rich, low-leverage compounders (BS score ≥ 6/10).
- Macro catalyst identification: every ENTER must cite a named, dated catalyst with a quantified magnitude.
- Technical confirmation: entry only when RSI, moving averages, and chart pattern align — cite exact levels.
- Position sizing is always $18,000 — no exceptions. You are either fully in or fully out. Conviction score is for transparency only and does not gate entry. If the setup is there, deploy full size. If you are not confident, skip entirely. Maximum {MAX_OPEN_POSITIONS} concurrent positions. Apply the winning pattern memory above — the goal is to compound on what works and avoid what has failed. Every trade you make teaches the system. Trade freely, size consistently, learn continuously.
- Stop loss: {STOP_LOSS_PCT*100:.0f}% below entry. Take profit: {TAKE_PROFIT_PCT*100:.0f}% above entry.
- RISK-OFF regime: only hedges (VXX, GLD, TLT, SH). No long equity.
- Sector overlap: do not duplicate sector exposure across open positions.
- Risk-reward floor: minimum 2.0:1 required on all new entries. Reject any trade below this.
- BANNED phrases: "strong fundamentals", "supportive macro", "well-positioned", "solid balance sheet" — every sentence must contain a specific number or named event.

{macro_block}

PORTFOLIO STATE:
{port_block}
{sig_block}

TRADE MEMORY:
{memory_to_prompt(memory)}

ANALYTICAL CHECKLIST — every ENTER action must satisfy ALL 8 points:
1. MACRO CATALYST: Name the specific event with probability or magnitude (e.g., "CME FedWatch 74% Sep pause", "PCE YoY printed +2.6% vs 2.9% prior").
2. COMPANY CATALYST: Exact EPS/revenue figure vs. estimate and beat % (e.g., "Q2 EPS $2.34 vs $2.10 est, +11.4% beat").
3. TECHNICAL: Exact RSI value, named chart pattern, 50-day MA price, 200-day MA price, volume context.
4. BALANCE SHEET: Exact cash ($XB), total debt ($XB), D/E ratio to 2 decimal places, retained earnings 3-year trend ($XB / $YB / $ZB), buybacks Y/N.
5. PRICE LEVELS: entry_price, stop_price (entry × {1-STOP_LOSS_PCT:.2f}), target_price (entry × {1+TAKE_PROFIT_PCT:.2f}) — all as specific dollar values.
6. RISK-REWARD: (target − entry) ÷ (entry − stop) ≥ 2.0. Compute and state the number.
7. SECTOR: Name the sector, portfolio weight after trade vs. S&P 500 sector weight, flag if >2× index weight.
8. DOWNSIDE SCENARIO: One specific quantified risk (e.g., "If PCE re-accelerates above 3.2%, growth multiple compression 15–20% implies $X.XX downside").

RESPOND ONLY with valid JSON — no preamble, no markdown fences:
{{
  "market_assessment": "2–3 sentences citing specific indicators: exact VIX level, yield curve spread in bps, HY spread in bps, and named macro data point with figure.",
  "regime": "RISK-ON|RISK-OFF|NEUTRAL|CAUTION",
  "actions": [
    {{
      "ticker": "XXXX",
      "action": "ENTER|EXIT|HOLD",
      "position_size_usd": 18000,
      "conviction": 8,
      "entry_price": 0.00,
      "stop_price": 0.00,
      "target_price": 0.00,
      "risk_reward_ratio": 0.0,
      "thesis": "7 sentences numbered (1)–(7). STRICT RULE: every sentence must contain at least one of: a price level ($X.XX), a percentage (X.X%), a ratio (X.Xx), or a named date/event. Any sentence without a number is invalid — rewrite it until it has one. Write as a PM who owns this position: concise, specific, zero filler, no generic descriptors. (1) Macro: exact VIX level and the specific named policy or data print driving the regime (e.g. 'PCE +2.6% vs 2.9% prior, CME FedWatch 74% Sep pause'). (2) Company catalyst: exact EPS or revenue figure, beat % vs estimate, and the event date. (3) Entry trigger: exact RSI value, pattern name, and the specific price that confirmed the entry. (4) MA confirmation: 50-day at $X.XX and 200-day at $Y.YY — state the spread as a % and what the cross/gap signals. (5) Balance sheet: cash $X.XB vs total debt $Y.YB, D/E ratio to 2 decimal places, retained earnings grew/fell X% last year. (6) Sizing: $X deployed = X.X% of the $183k portfolio, leaves $Y cash, room for N more full-size positions. (7) Stop and target: stop at $X.XX (-8%), target $Y.YY (+20%), R:R = Z.Z:1 — name the exact catalyst or price level that would trigger the stop.",
      "catalyst": "Single named catalyst with exact figure, date, and why it is a near-term price driver.",
      "technical_setup": "RSI X.X, pattern name, price vs. 50-day $X.XX and 200-day $Y.YY, volume X% above 20-day avg.",
      "balance_sheet_read": "Cash $X.XB, total debt $Y.YB, D/E Z.ZZ, retained earnings $AB/$BB/$CB (3yr most-recent-first), buybacks Y/N, preferred Y/N, BS score N/10.",
      "sector_positioning": "Sector name. Portfolio sector weight after trade X.X% vs. SPX weight Y.Y%. Concentration flag: Y/N.",
      "risk_factors": "Named risk with specific quantified scenario and the exact dollar stop that limits the loss.",
      "tier": "1|2|3|4"
    }}
  ],
  "portfolio_notes": "Cash reserve %, sector concentration summary with weights, and overall risk posture — all with specific numbers.",
  "memory_applied": "Which past lessons influenced today — cite specific ticker and date if applicable."
}}"""

def call_claude(prompt: str) -> Optional[dict]:
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    try:
        log.info("Calling Claude Opus 4.8...")
        msg = client.messages.create(
            model="claude-opus-4-8",
            max_tokens=6000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = msg.content[0].text.strip()
        # Strip any accidental markdown fences
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        decision = json.loads(raw.strip())
        log.info(f"Decision: regime={decision.get('regime')} | actions={len(decision.get('actions',[]))}")
        return decision
    except json.JSONDecodeError as e:
        log.error(f"JSON parse error: {e}")
    except Exception as e:
        log.error(f"Claude API error: {e}")
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 7. TRADE EXECUTOR
# ─────────────────────────────────────────────────────────────────────────────
def execute(decision: dict, portfolio: dict, trade_log: list) -> tuple:
    actions_taken = []
    held_tickers  = {p["ticker"] for p in portfolio.get("positions", [])}

    for action in decision.get("actions", []):
        ticker     = action.get("ticker", "").upper().strip()
        act        = action.get("action", "HOLD").upper()
        size_usd   = MAX_POSITION_USD        # flat $18k — no conviction-based sizing
        conviction = int(action.get("conviction", 5))

        if not ticker:
            continue

        if act == "ENTER":
            if ticker in held_tickers:
                log.info(f"SKIP {ticker}: already held.")
                continue
            if len(portfolio["positions"]) >= MAX_OPEN_POSITIONS:
                log.warning(f"SKIP {ticker}: max positions reached.")
                continue
            if portfolio["cash"] < size_usd:
                log.warning(f"SKIP {ticker}: insufficient cash.")
                continue
            if conviction < MIN_CONVICTION:
                log.info(f"SKIP {ticker}: conviction {conviction} < {MIN_CONVICTION}.")
                continue

            price = fetch_price(ticker)
            if not price:
                log.error(f"SKIP {ticker}: price unavailable.")
                continue

            shares   = size_usd / price
            position = {
                "ticker":           ticker,
                "shares":           shares,
                "entry_price":      price,
                "cost_basis_total": size_usd,
                "current_price":    price,
                "current_value":    size_usd,
                "unrealised_pnl":   0.0,
                "unrealised_pct":   0.0,
                "thesis_summary":    action.get("thesis", ""),
                "catalyst":         action.get("catalyst", ""),
                "technical_setup":  action.get("technical_setup", ""),
                "risk_factors":     action.get("risk_factors", ""),
                "balance_sheet_read": action.get("balance_sheet_read", ""),
                "conviction":       conviction,
                "tier":             action.get("tier", "2"),
                "entry_date":       datetime.utcnow().isoformat(),
            }
            portfolio["positions"].append(position)
            portfolio["cash"] -= size_usd
            held_tickers.add(ticker)

            trade_log.append({**position, "type": "ENTER",
                               "trade_date": datetime.utcnow().isoformat()})
            actions_taken.append(action)
            log.info(f"ENTER {ticker}: {shares:.2f}sh @ ${price:.2f} | Conv {conviction}/10")

        elif act == "EXIT":
            for i, pos in enumerate(portfolio["positions"]):
                if pos["ticker"] == ticker:
                    price = fetch_price(ticker) or pos.get("current_price", pos["entry_price"])
                    pos.update({
                        "current_price":  price,
                        "current_value":  price * pos["shares"],
                        "unrealised_pnl": (price * pos["shares"]) - pos["cost_basis_total"],
                        "unrealised_pct": ((price * pos["shares"]) - pos["cost_basis_total"]) / pos["cost_basis_total"],
                    })
                    t = close_pos(pos, action.get("thesis", "AI exit decision"))
                    t["type"] = "EXIT"
                    trade_log.append(t)
                    portfolio["cash"] += pos["current_value"]
                    portfolio["positions"].pop(i)
                    held_tickers.discard(ticker)
                    actions_taken.append(action)
                    log.info(f"EXIT {ticker}: P&L {pos['unrealised_pct']*100:+.1f}%")
                    break

    return portfolio, trade_log, actions_taken


# ─────────────────────────────────────────────────────────────────────────────
# 8. PDF REPORT  (Goldman Sachs morning note style)
# ─────────────────────────────────────────────────────────────────────────────
NAVY  = colors.HexColor("#0d1f3c")
BLUE  = colors.HexColor("#2a6db5")
GREEN = colors.HexColor("#1a6e3e")
RED   = colors.HexColor("#9e2020")
LIGHT = colors.HexColor("#f4f8fd")
AMBER = colors.HexColor("#d07a10")
RULE  = colors.HexColor("#dde5f0")
BODY_TEXT = colors.HexColor("#2c3e50")


def _parse_thesis_points(thesis: str) -> list:
    """Split '(1) ... (2) ... (3) ...' into individual point strings."""
    parts = re.split(r'\s*\((\d+)\)\s*', thesis.strip())
    # parts: ['pre-text', '1', 'sentence1', '2', 'sentence2', ...]
    result = []
    i = 1
    while i < len(parts) - 1:
        num  = parts[i]
        text = parts[i + 1].strip()
        if text:
            result.append(f"({num})  {text}")
        i += 2
    return result if result else []


def build_pdf(decision: dict, portfolio: dict, macro: dict,
              actions: list, auto_exits: list) -> bytes:
    """Goldman Sachs-style institutional trade report."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.65*inch, rightMargin=0.65*inch,
                            topMargin=0.6*inch, bottomMargin=0.6*inch)

    # ── Colour palette ────────────────────────────────────────────────────────
    NAVY  = colors.HexColor("#0d1f3c")
    BLUE  = colors.HexColor("#1a4a8a")
    BLUE2 = colors.HexColor("#2a6db5")
    GOLD  = colors.HexColor("#b8962e")
    GREEN = colors.HexColor("#1a6e3e")
    RED   = colors.HexColor("#9e2020")
    AMBER = colors.HexColor("#b86e00")
    LIGHT = colors.HexColor("#f0f4f9")
    CARD  = colors.HexColor("#f8faff")
    GREY  = colors.HexColor("#6b7280")
    WHITE = colors.white
    BLACK = colors.HexColor("#111827")

    # ── Style factory ─────────────────────────────────────────────────────────
    SS = getSampleStyleSheet()
    def S(name, **kw):
        return ParagraphStyle(name, parent=SS["Normal"], **kw)

    body_s = S("body", fontName="Helvetica", fontSize=8.5, textColor=BLACK,
               leading=13, spaceAfter=0, alignment=TA_JUSTIFY)

    story    = []
    now_str  = datetime.utcnow().strftime("%d %b %Y · %H:%M UTC")
    window   = "PRE-CLOSE · 3:30 PM ET" if datetime.utcnow().hour >= 19 else "OPEN · 9:35 AM ET"
    regime   = decision.get("regime", macro.get("regime", "NEUTRAL"))
    rc_map   = {"RISK-ON": GREEN, "RISK-OFF": RED, "CAUTION": AMBER}
    rc       = rc_map.get(regime, BLUE2)
    ind      = macro.get("indicators", {})
    port_val = portfolio.get("total_value", 0)
    cash     = portfolio.get("cash", 0)
    pnl      = port_val - 183_000
    pnl_pct  = (pnl / 183_000) * 100
    pos_list = portfolio.get("positions", [])

    def iv(k):
        v = ind.get(k, {}).get("value")
        return f"{v:.2f}" if v is not None else "—"

    # ── COVER HEADER ──────────────────────────────────────────────────────────
    header_data = [[
        Paragraph("MeritQuant", S("ht", fontName="Helvetica-Bold", fontSize=22,
                                   textColor=WHITE, spaceAfter=0)),
        Paragraph(f"Autonomous Trade Report<br/><font size='9' color='#a0b8d4'>{window} · {now_str}</font>",
                  S("hs", fontName="Helvetica", fontSize=13, textColor=WHITE,
                    spaceAfter=0, leading=18)),
    ]]
    ht = Table(header_data, colWidths=[2.8*inch, 4.6*inch])
    ht.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), NAVY),
        ("TOPPADDING",    (0,0), (-1,-1), 18),
        ("BOTTOMPADDING", (0,0), (-1,-1), 18),
        ("LEFTPADDING",   (0,0), (0,-1),  16),
        ("LEFTPADDING",   (1,0), (1,-1),  10),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(ht)
    story.append(Table([[""]], colWidths=[7.45*inch],
                       style=[("BACKGROUND",    (0,0), (-1,-1), GOLD),
                              ("TOPPADDING",    (0,0), (-1,-1), 1),
                              ("BOTTOMPADDING", (0,0), (-1,-1), 1)]))
    story.append(Spacer(1, 10))

    # ── KPI ROW ───────────────────────────────────────────────────────────────
    pnl_col  = GREEN if pnl >= 0 else RED
    cash_pct = (cash / port_val * 100) if port_val else 0
    kpi_vals = [
        ("PORTFOLIO VALUE", f"${port_val:,.0f}", NAVY),
        ("NET P&L vs COST", f"{'+'if pnl>=0 else ''}{pnl:,.0f} ({pnl_pct:+.2f}%)", pnl_col),
        ("CASH RESERVE",    f"${cash:,.0f} ({cash_pct:.1f}%)", BLUE),
        ("POSITIONS",       f"{len(pos_list)} / 6", NAVY),
        ("REGIME",          regime, rc),
    ]
    kpi_cells = []
    for lbl, v, col in kpi_vals:
        kpi_cells.append(
            Table([[Paragraph(lbl, S(f"kl{lbl}", fontName="Helvetica-Bold", fontSize=6.5,
                                     textColor=GREY, letterSpacing=1, spaceAfter=2))],
                   [Paragraph(v,   S(f"kv{lbl}", fontName="Helvetica-Bold", fontSize=11,
                                     textColor=col, spaceAfter=0))]],
                  colWidths=[1.38*inch],
                  style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                         ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#dde8f5")),
                         ("TOPPADDING",    (0,0), (-1,-1), 9),
                         ("BOTTOMPADDING", (0,0), (-1,-1), 9),
                         ("LEFTPADDING",   (0,0), (-1,-1), 10)])
        )
    kpi_row = Table([kpi_cells], colWidths=[1.49*inch]*5)
    kpi_row.setStyle(TableStyle([
        ("TOPPADDING",    (0,0), (-1,-1), 0), ("BOTTOMPADDING", (0,0), (-1,-1), 0),
        ("LEFTPADDING",   (0,0), (-1,-1), 0), ("RIGHTPADDING",  (0,0), (-1,-1), 4),
    ]))
    story.append(kpi_row)
    story.append(Spacer(1, 12))

    # ── MARKET ASSESSMENT ─────────────────────────────────────────────────────
    assess_text = decision.get("market_assessment", "—")
    assess_box  = Table([[
        Paragraph("MARKET ASSESSMENT", S("mah", fontName="Helvetica-Bold", fontSize=7,
                                          textColor=BLUE2, letterSpacing=2, spaceAfter=4)),
        Paragraph(regime, S("reg", fontName="Helvetica-Bold", fontSize=8,
                             textColor=rc, spaceAfter=0)),
    ],[
        Paragraph(assess_text, body_s), "",
    ]], colWidths=[5.8*inch, 1.5*inch])
    assess_box.setStyle(TableStyle([
        ("SPAN",          (0,1), (1,1)),
        ("BACKGROUND",    (0,0), (-1,-1), CARD),
        ("BOX",           (0,0), (-1,-1), 0.8, BLUE2),
        ("LINEBEFORE",    (0,0), (0,-1),  3, BLUE2),
        ("TOPPADDING",    (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING",   (0,0), (-1,-1), 12), ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
    ]))
    story.append(assess_box)
    story.append(Spacer(1, 10))

    # ── MACRO SNAPSHOT ────────────────────────────────────────────────────────
    macro_hdr = Table([["INDICATOR", "VALUE", "INDICATOR", "VALUE"]],
                      colWidths=[1.95*inch, 0.8*inch, 1.95*inch, 0.8*inch])
    macro_hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), NAVY), ("TEXTCOLOR", (0,0), (-1,-1), WHITE),
        ("FONTNAME",      (0,0), (-1,-1), "Helvetica-Bold"), ("FONTSIZE", (0,0), (-1,-1), 7),
        ("TOPPADDING",    (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
    ]))
    macro_rows = [
        ["VIX",                    iv("VIX"),    "HY Credit Spread (bps)", iv("HY_SPREAD")],
        ["Yield Curve 10Y-2Y (%)", iv("T10Y2Y"), "10Y Treasury (%)",       iv("T10Y")],
        ["US Dollar (DXY)",        iv("DXY"),    "Fed Funds Rate (%)",     iv("FED_RATE")],
    ]
    macro_body = Table(macro_rows, colWidths=[1.95*inch, 0.8*inch, 1.95*inch, 0.8*inch])
    macro_body.setStyle(TableStyle([
        ("FONTNAME",       (0,0), (-1,-1), "Helvetica"), ("FONTSIZE", (0,0), (-1,-1), 8),
        ("FONTNAME",       (1,0), (1,-1),  "Helvetica-Bold"),
        ("FONTNAME",       (3,0), (3,-1),  "Helvetica-Bold"),
        ("TEXTCOLOR",      (1,0), (1,-1),  NAVY), ("TEXTCOLOR", (3,0), (3,-1), NAVY),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [WHITE, CARD]),
        ("GRID",           (0,0), (-1,-1), 0.4, colors.HexColor("#dde8f5")),
        ("TOPPADDING",     (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",    (0,0), (-1,-1), 8),
    ]))
    story.append(macro_hdr)
    story.append(macro_body)
    story.append(Spacer(1, 12))

    # ── TRADE DECISIONS ───────────────────────────────────────────────────────
    if actions:
        story.append(Table([["TRADE DECISIONS"]], colWidths=[7.45*inch],
                           style=[("BACKGROUND",    (0,0), (-1,-1), NAVY),
                                  ("TEXTCOLOR",     (0,0), (-1,-1), WHITE),
                                  ("FONTNAME",      (0,0), (-1,-1), "Helvetica-Bold"),
                                  ("FONTSIZE",      (0,0), (-1,-1), 8),
                                  ("TOPPADDING",    (0,0), (-1,-1), 7),
                                  ("BOTTOMPADDING", (0,0), (-1,-1), 7),
                                  ("LEFTPADDING",   (0,0), (-1,-1), 12)]))
        story.append(Spacer(1, 4))

        for a in actions:
            act    = a.get("action", "")
            ticker = a.get("ticker", "")
            size   = a.get("position_size_usd", 0)
            conv   = a.get("conviction", 0)
            tier   = a.get("tier", "?")
            thesis = a.get("thesis", "—")
            cat    = a.get("catalyst", "—")
            tech   = a.get("technical_setup", "—")
            risk   = a.get("risk_factors", "—")
            bs     = a.get("balance_sheet_read", "—")
            entry  = a.get("entry_price", 0)
            stop   = a.get("stop_price", 0)
            target = a.get("target_price", 0)
            rr     = a.get("risk_reward_ratio", "—")
            ac     = GREEN if act == "ENTER" else RED if act == "EXIT" else BLUE2

            # Action banner
            story.append(Table([[
                Paragraph(f"{act} — {ticker}", S("ab", fontName="Helvetica-Bold",
                                                   fontSize=13, textColor=WHITE, spaceAfter=0)),
                Paragraph(f"${size:,.0f} · Conviction {conv}/10 · Tier {tier}",
                          S("am", fontName="Helvetica", fontSize=9,
                            textColor=WHITE, spaceAfter=0, alignment=TA_RIGHT)),
            ]], colWidths=[4*inch, 3.45*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), ac),
                       ("TOPPADDING",    (0,0), (-1,-1), 9), ("BOTTOMPADDING", (0,0), (-1,-1), 9),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12), ("RIGHTPADDING",  (0,0), (-1,-1), 12),
                       ("VALIGN",        (0,0), (-1,-1), "MIDDLE")]))

            # Price row
            if any([entry, stop, target]):
                pr_cells = []
                for lbl, pval, col in [
                    ("ENTRY",         f"${entry:.2f}" if entry else "—",  NAVY),
                    ("STOP LOSS",     f"${stop:.2f}"  if stop  else "—",  RED),
                    ("PRICE TARGET",  f"${target:.2f}" if target else "—", GREEN),
                    ("RISK : REWARD", str(rr), BLUE2),
                ]:
                    pr_cells.append(Table([[
                        Paragraph(lbl, S(f"prl{lbl}", fontName="Helvetica-Bold",
                                          fontSize=6.5, textColor=GREY, letterSpacing=1, spaceAfter=2))
                    ],[
                        Paragraph(pval, S(f"prv{lbl}", fontName="Helvetica-Bold",
                                           fontSize=11, textColor=col, spaceAfter=0))
                    ]], colWidths=[1.85*inch],
                        style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                               ("TOPPADDING",    (0,0), (-1,-1), 7),
                               ("BOTTOMPADDING", (0,0), (-1,-1), 7),
                               ("LEFTPADDING",   (0,0), (-1,-1), 10),
                               ("BOX",           (0,0), (-1,-1), 0.4, colors.HexColor("#dde8f5"))]))
                story.append(Table([pr_cells], colWidths=[1.865*inch]*4,
                                   style=[("TOPPADDING",    (0,0), (-1,-1), 0),
                                          ("BOTTOMPADDING", (0,0), (-1,-1), 0),
                                          ("LEFTPADDING",   (0,0), (-1,-1), 0),
                                          ("RIGHTPADDING",  (0,0), (-1,-1), 0)]))

            # Section block helper
            def section_block(label, text, border_col=BLUE2):
                return Table([[
                    Paragraph(label, S(f"sl{label}", fontName="Helvetica-Bold", fontSize=7,
                                       textColor=border_col, letterSpacing=1.5, spaceAfter=3))
                ],[
                    Paragraph(text, body_s),
                ]], colWidths=[7.45*inch],
                    style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                           ("LINEBEFORE",    (0,0), (0,-1),  3, border_col),
                           ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                           ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                           ("LEFTPADDING",   (0,0), (-1,-1), 12), ("RIGHTPADDING",  (0,0), (-1,-1), 12)])

            story.append(Spacer(1, 2))
            story.append(section_block("INVESTMENT THESIS", thesis, BLUE2))
            story.append(Spacer(1, 2))
            story.append(section_block("PRIMARY CATALYST", cat, GOLD))
            story.append(Spacer(1, 2))

            # Balance sheet + Technical side by side
            bs_block = Table([[
                Paragraph("BALANCE SHEET", S("bsl", fontName="Helvetica-Bold", fontSize=7,
                                              textColor=GREEN, letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(bs, S("bsb", fontName="Helvetica", fontSize=8.5,
                                 textColor=BLACK, leading=13, spaceAfter=0))
            ]], colWidths=[3.6*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, GREEN),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12)])
            tech_block = Table([[
                Paragraph("TECHNICAL SETUP", S("tsl", fontName="Helvetica-Bold", fontSize=7,
                                               textColor=AMBER, letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(tech, S("tsb", fontName="Helvetica", fontSize=8.5,
                                   textColor=BLACK, leading=13, spaceAfter=0))
            ]], colWidths=[3.65*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, AMBER),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12)])
            story.append(Table([[bs_block, tech_block]],
                               colWidths=[3.72*inch, 3.73*inch],
                               style=[("TOPPADDING",    (0,0), (-1,-1), 0),
                                      ("BOTTOMPADDING", (0,0), (-1,-1), 0),
                                      ("LEFTPADDING",   (0,0), (-1,-1), 0),
                                      ("RIGHTPADDING",  (0,0), (0,-1),  4),
                                      ("RIGHTPADDING",  (1,0), (1,-1),  0)]))
            story.append(Spacer(1, 2))
            story.append(section_block("KEY RISK FACTORS", risk, RED))
            story.append(Spacer(1, 10))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dde8f5")))
            story.append(Spacer(1, 8))

    elif not pos_list:
        story.append(Table([["NO TRADE DECISIONS THIS SESSION"]], colWidths=[7.45*inch],
                           style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                                  ("TEXTCOLOR",     (0,0), (-1,-1), GREY),
                                  ("FONTNAME",      (0,0), (-1,-1), "Helvetica"),
                                  ("FONTSIZE",      (0,0), (-1,-1), 9),
                                  ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                                  ("TOPPADDING",    (0,0), (-1,-1), 18),
                                  ("BOTTOMPADDING", (0,0), (-1,-1), 18)]))
    story.append(Spacer(1, 10))

    # ── OPEN POSITIONS — full per-position cards ──────────────────────────────
    if pos_list:
        story.append(Table([["OPEN POSITIONS — FULL POSITION REVIEW"]],
                           colWidths=[7.45*inch],
                           style=[("BACKGROUND",    (0,0), (-1,-1), NAVY),
                                  ("TEXTCOLOR",     (0,0), (-1,-1), WHITE),
                                  ("FONTNAME",      (0,0), (-1,-1), "Helvetica-Bold"),
                                  ("FONTSIZE",      (0,0), (-1,-1), 8),
                                  ("TOPPADDING",    (0,0), (-1,-1), 7),
                                  ("BOTTOMPADDING", (0,0), (-1,-1), 7),
                                  ("LEFTPADDING",   (0,0), (-1,-1), 12)]))
        story.append(Spacer(1, 6))

        for pos in pos_list:
            tk      = pos.get("ticker", "")
            shares  = pos.get("shares", 0)
            ep      = pos.get("entry_price", 0)
            cp      = pos.get("current_price", ep)
            cv      = pos.get("current_value", 0)
            pu      = pos.get("unrealised_pnl", 0)
            pp      = pos.get("unrealised_pct", 0)
            thesis  = pos.get("thesis_summary", "No thesis recorded.")
            cat     = pos.get("catalyst", "—")
            tech    = pos.get("technical_setup", "—")
            risk    = pos.get("risk_factors", "—")
            conv    = pos.get("conviction", "—")
            tier    = pos.get("tier", "—")
            bs_read = pos.get("balance_sheet_read", "—")
            entry_d = pos.get("entry_date", "")[:10]
            pc      = GREEN if pp >= 0 else RED

            # Position header
            story.append(Table([[
                Paragraph(f"{tk}", S(f"ph{tk}", fontName="Helvetica-Bold",
                                     fontSize=16, textColor=WHITE, spaceAfter=0)),
                Paragraph(
                    f"<font size='8' color='#a0b8d4'>{shares:.1f} shares · Entry ${ep:.2f}"
                    f" · Conv {conv}/10 · Tier {tier} · Since {entry_d}</font>",
                    S(f"pm{tk}", fontName="Helvetica", fontSize=8, textColor=WHITE,
                      spaceAfter=0, alignment=TA_RIGHT)),
            ]], colWidths=[3*inch, 4.45*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), BLUE),
                       ("TOPPADDING",    (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10),
                       ("LEFTPADDING",   (0,0), (-1,-1), 14), ("RIGHTPADDING",  (0,0), (-1,-1), 14),
                       ("VALIGN",        (0,0), (-1,-1), "MIDDLE")]))

            # P&L metrics row
            pm_cells = []
            for lbl, pval, col in [
                ("CURRENT PRICE",  f"${cp:.2f}", NAVY),
                ("POSITION VALUE", f"${cv:,.0f}", NAVY),
                ("UNREALISED P&L", f"{'+'if pu>=0 else ''}${pu:,.0f}", pc),
                ("RETURN",         f"{pp*100:+.2f}%", pc),
            ]:
                pm_cells.append(Table([[
                    Paragraph(lbl, S(f"pml{lbl}{tk}", fontName="Helvetica-Bold",
                                     fontSize=6.5, textColor=GREY,
                                     letterSpacing=1, spaceAfter=2))
                ],[
                    Paragraph(pval, S(f"pmv{lbl}{tk}", fontName="Helvetica-Bold",
                                      fontSize=11, textColor=col, spaceAfter=0))
                ]], colWidths=[1.85*inch],
                    style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                           ("TOPPADDING",    (0,0), (-1,-1), 7),
                           ("BOTTOMPADDING", (0,0), (-1,-1), 7),
                           ("LEFTPADDING",   (0,0), (-1,-1), 10),
                           ("BOX",           (0,0), (-1,-1), 0.4, colors.HexColor("#dde8f5"))]))
            story.append(Table([pm_cells], colWidths=[1.865*inch]*4,
                               style=[("TOPPADDING",    (0,0), (-1,-1), 0),
                                      ("BOTTOMPADDING", (0,0), (-1,-1), 0),
                                      ("LEFTPADDING",   (0,0), (-1,-1), 0),
                                      ("RIGHTPADDING",  (0,0), (-1,-1), 0)]))

            # Thesis
            story.append(Table([[
                Paragraph("INVESTMENT THESIS", S(f"ptl{tk}", fontName="Helvetica-Bold",
                                                  fontSize=7, textColor=BLUE2,
                                                  letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(thesis, S(f"ptb{tk}", fontName="Helvetica", fontSize=8.5,
                                    textColor=BLACK, leading=13.5, spaceAfter=0,
                                    alignment=TA_JUSTIFY))
            ]], colWidths=[7.45*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, BLUE2),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12), ("RIGHTPADDING",  (0,0), (-1,-1), 12)]))
            story.append(Spacer(1, 2))

            # Catalyst + Technical side by side
            cat_blk = Table([[
                Paragraph("CATALYST", S(f"catl{tk}", fontName="Helvetica-Bold", fontSize=7,
                                        textColor=GOLD, letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(cat, S(f"catb{tk}", fontName="Helvetica", fontSize=8.5,
                                  textColor=BLACK, leading=13, spaceAfter=0))
            ]], colWidths=[3.6*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, GOLD),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12)])
            tch_blk = Table([[
                Paragraph("TECHNICAL SETUP", S(f"techl{tk}", fontName="Helvetica-Bold",
                                               fontSize=7, textColor=AMBER,
                                               letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(tech, S(f"techb{tk}", fontName="Helvetica", fontSize=8.5,
                                   textColor=BLACK, leading=13, spaceAfter=0))
            ]], colWidths=[3.65*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, AMBER),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12)])
            story.append(Table([[cat_blk, tch_blk]],
                               colWidths=[3.72*inch, 3.73*inch],
                               style=[("TOPPADDING",    (0,0), (-1,-1), 0),
                                      ("BOTTOMPADDING", (0,0), (-1,-1), 0),
                                      ("LEFTPADDING",   (0,0), (-1,-1), 0),
                                      ("RIGHTPADDING",  (0,0), (0,-1),  4),
                                      ("RIGHTPADDING",  (1,0), (1,-1),  0)]))
            story.append(Spacer(1, 2))

            # Balance sheet + Risk side by side
            bs_blk = Table([[
                Paragraph("BALANCE SHEET", S(f"bsl2{tk}", fontName="Helvetica-Bold",
                                              fontSize=7, textColor=GREEN,
                                              letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(bs_read if bs_read else "—",
                          S(f"bsb2{tk}", fontName="Helvetica", fontSize=8.5,
                            textColor=BLACK, leading=13, spaceAfter=0))
            ]], colWidths=[3.6*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, GREEN),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12)])
            rsk_blk = Table([[
                Paragraph("KEY RISKS", S(f"rskl{tk}", fontName="Helvetica-Bold",
                                         fontSize=7, textColor=RED,
                                         letterSpacing=1.5, spaceAfter=3))
            ],[
                Paragraph(risk if risk != "—" else "No specific risk factors recorded.",
                          S(f"rskb{tk}", fontName="Helvetica", fontSize=8.5,
                            textColor=BLACK, leading=13, spaceAfter=0))
            ]], colWidths=[3.65*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("LINEBEFORE",    (0,0), (0,-1),  3, RED),
                       ("TOPPADDING",    (0,0), (0, 0),  8), ("BOTTOMPADDING", (0,0), (0,0), 2),
                       ("TOPPADDING",    (0,1), (0, 1),  4), ("BOTTOMPADDING", (0,1), (0,1), 8),
                       ("LEFTPADDING",   (0,0), (-1,-1), 12)])
            story.append(Table([[bs_blk, rsk_blk]],
                               colWidths=[3.72*inch, 3.73*inch],
                               style=[("TOPPADDING",    (0,0), (-1,-1), 0),
                                      ("BOTTOMPADDING", (0,0), (-1,-1), 0),
                                      ("LEFTPADDING",   (0,0), (-1,-1), 0),
                                      ("RIGHTPADDING",  (0,0), (0,-1),  4),
                                      ("RIGHTPADDING",  (1,0), (1,-1),  0)]))
            story.append(Spacer(1, 12))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dde8f5")))
            story.append(Spacer(1, 8))

    # ── AUTO EXITS ────────────────────────────────────────────────────────────
    if auto_exits:
        story.append(Spacer(1, 4))
        story.append(Table([["AUTOMATIC EXITS (STOP LOSS / TAKE PROFIT)"]],
                           colWidths=[7.45*inch],
                           style=[("BACKGROUND",    (0,0), (-1,-1), RED),
                                  ("TEXTCOLOR",     (0,0), (-1,-1), WHITE),
                                  ("FONTNAME",      (0,0), (-1,-1), "Helvetica-Bold"),
                                  ("FONTSIZE",      (0,0), (-1,-1), 8),
                                  ("TOPPADDING",    (0,0), (-1,-1), 7),
                                  ("BOTTOMPADDING", (0,0), (-1,-1), 7),
                                  ("LEFTPADDING",   (0,0), (-1,-1), 12)]))
        story.append(Spacer(1, 4))
        for ticker, reason, pnl_p in auto_exits:
            col = GREEN if pnl_p > 0 else RED
            story.append(Table([[
                Paragraph(f"{ticker} — {reason}",
                          S(f"ex{ticker}", fontName="Helvetica-Bold",
                            fontSize=10, textColor=col, spaceAfter=0)),
                Paragraph(f"P&L: {pnl_p*100:+.1f}%",
                          S(f"exp{ticker}", fontName="Helvetica-Bold",
                            fontSize=11, textColor=col, spaceAfter=0,
                            alignment=TA_RIGHT)),
            ]], colWidths=[5*inch, 2.45*inch],
                style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                       ("BOX",           (0,0), (-1,-1), 0.5, col),
                       ("TOPPADDING",    (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10),
                       ("LEFTPADDING",   (0,0), (-1,-1), 14), ("RIGHTPADDING",  (0,0), (-1,-1), 14),
                       ("VALIGN",        (0,0), (-1,-1), "MIDDLE")]))
            story.append(Spacer(1, 4))

    # ── PORTFOLIO NOTES + MEMORY ──────────────────────────────────────────────
    notes  = decision.get("portfolio_notes", "")
    memory = decision.get("memory_applied", "")
    if notes:
        story.append(Spacer(1, 4))
        story.append(Table([[
            Paragraph("PORTFOLIO NOTES", S("pnl", fontName="Helvetica-Bold", fontSize=7,
                                           textColor=NAVY, letterSpacing=1.5, spaceAfter=4))
        ],[
            Paragraph(notes, body_s),
        ]], colWidths=[7.45*inch],
            style=[("BACKGROUND",    (0,0), (-1,-1), LIGHT),
                   ("BOX",           (0,0), (-1,-1), 0.8, NAVY),
                   ("TOPPADDING",    (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10),
                   ("LEFTPADDING",   (0,0), (-1,-1), 12), ("RIGHTPADDING",  (0,0), (-1,-1), 12)]))
        story.append(Spacer(1, 6))
    if memory:
        story.append(Table([[
            Paragraph("MEMORY APPLIED", S("mapl", fontName="Helvetica-Bold", fontSize=7,
                                          textColor=BLUE2, letterSpacing=1.5, spaceAfter=4))
        ],[
            Paragraph(memory, body_s),
        ]], colWidths=[7.45*inch],
            style=[("BACKGROUND",    (0,0), (-1,-1), CARD),
                   ("LINEBEFORE",    (0,0), (0,-1),  3, GOLD),
                   ("TOPPADDING",    (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10),
                   ("LEFTPADDING",   (0,0), (-1,-1), 12), ("RIGHTPADDING",  (0,0), (-1,-1), 12)]))

    # ── FOOTER ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 16))
    story.append(Table([["MeritQuant Autonomous Trader · Paper Portfolio Mirror · Not financial advice"]],
                       colWidths=[7.45*inch],
                       style=[("BACKGROUND",    (0,0), (-1,-1), NAVY),
                              ("TEXTCOLOR",     (0,0), (-1,-1), colors.HexColor("#6b8cb0")),
                              ("FONTNAME",      (0,0), (-1,-1), "Helvetica"),
                              ("FONTSIZE",      (0,0), (-1,-1), 7),
                              ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                              ("TOPPADDING",    (0,0), (-1,-1), 8),
                              ("BOTTOMPADDING", (0,0), (-1,-1), 8)]))
    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 9. EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(decision: dict, portfolio: dict, trade_log: list) -> bytes:
    wb   = openpyxl.Workbook()
    NAVY_H, BLUE_H, GREEN_H, RED_H, LIGHT_H = "0D1F3C","2A6DB5","1A6E3E","9E2020","F4F8FD"

    def hdr(ws, row, ncols, text, fill=None):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = PatternFill("solid", fgColor=fill or NAVY_H)
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    hdr(ws, 1, 2, "MeritQuant — AI Trade Session Report")
    rows = [
        ("Generated UTC", datetime.utcnow().strftime("%Y-%m-%d %H:%M")),
        ("Portfolio Value", f"${portfolio.get('total_value',0):,.2f}"),
        ("Cash Available", f"${portfolio.get('cash',0):,.2f}"),
        ("Open Positions", f"{len(portfolio.get('positions',[]))}/{MAX_OPEN_POSITIONS}"),
        ("Regime", decision.get("regime","?")),
        ("Market Assessment", decision.get("market_assessment","—")),
        ("Memory Applied", decision.get("memory_applied","—")),
        ("Portfolio Notes", decision.get("portfolio_notes","—")),
    ]
    for r, (k, v) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=9)
        ws.cell(row=r, column=2, value=v).font = Font(size=9)
        if r % 2 == 0:
            for c in range(1, 3):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_H)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 55

    # ── Sheet 2: Positions ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Positions")
    cols = ["Ticker","Shares","Entry $","Current $","Value $","P&L $","P&L %",
            "Conviction","Tier","Catalyst","Technical Setup","Entry Date"]
    for c, h in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor=NAVY_H)
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        ws2.column_dimensions[cell.column_letter].width = 16
    for r, p in enumerate(portfolio.get("positions",[]), 2):
        pnl = p.get("unrealised_pct", 0)
        row_vals = [
            p["ticker"], round(p["shares"],2), round(p["entry_price"],2),
            round(p.get("current_price", p["entry_price"]),2),
            round(p.get("current_value",0),2), round(p.get("unrealised_pnl",0),2),
            f"{pnl*100:.1f}%", p.get("conviction","—"), p.get("tier","—"),
            p.get("catalyst",""), p.get("technical_setup",""),
            p.get("entry_date","")[:10]
        ]
        for c, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(size=9,
                             color=(GREEN_H if pnl >= 0 else RED_H) if c == 7 else "000000",
                             bold=(c == 7))
        if r % 2 == 0:
            for c in range(1, len(cols)+1):
                ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT_H)

    # ── Sheet 3: Trade Log ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Trade Log")
    log_cols = ["Date","Type","Ticker","Shares","Entry $","Exit $","P&L $","P&L %","Reason"]
    for c, h in enumerate(log_cols, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor=NAVY_H)
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        ws3.column_dimensions[cell.column_letter].width = 14
    for r, t in enumerate(reversed(trade_log[-100:]), 2):
        pnl = t.get("pnl_pct", 0)
        row_vals = [
            str(t.get("trade_date", t.get("entry_date","—")))[:10],
            t.get("type","—"), t.get("ticker","—"),
            round(t.get("shares",0),2), round(t.get("entry_price",0),2),
            round(t.get("exit_price",0),2), round(t.get("pnl_usd",0),2),
            f"{pnl*100:.1f}%" if pnl else "—",
            str(t.get("exit_reason", t.get("thesis_summary","")))[:80],
        ]
        for c, val in enumerate(row_vals, 1):
            ws3.cell(row=r, column=c, value=val).font = Font(size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 10. TELEGRAM MESSAGES
# ─────────────────────────────────────────────────────────────────────────────
def trade_msg(action: dict, portfolio: dict) -> str:
    act = action.get("action","")
    e   = {"ENTER":"🟢","EXIT":"🔴","HOLD":"⚪️"}.get(act,"⚪️")
    return (
        f"{e} <b>MeritQuant — {act} {action.get('ticker')}</b>\n\n"
        f"💰 Size: <b>${action.get('position_size_usd',0):,.0f}</b>\n"
        f"🎯 Conviction: <b>{action.get('conviction','—')}/10</b>  |  Tier {action.get('tier','?')}\n\n"
        f"📋 <b>Thesis:</b>\n{action.get('thesis','—')}\n\n"
        f"⚡️ <b>Catalyst:</b> {action.get('catalyst','—')}\n"
        f"📊 <b>Setup:</b> {action.get('technical_setup','—')}\n"
        f"⚠️ <b>Risk:</b> {action.get('risk_factors','—')}\n\n"
        f"💼 Portfolio: ${portfolio.get('total_value',0):,.0f}  |  Cash: ${portfolio.get('cash',0):,.0f}"
    )

def sl_tp_msg(ticker: str, reason: str, pnl: float) -> str:
    e = "✅" if pnl > 0 else "🛑"
    return f"{e} <b>MeritQuant — AUTO EXIT {ticker}</b>\n📋 {reason}\n💰 P&L: {pnl*100:+.1f}%"

def session_msg(decision: dict, portfolio: dict, actions: list,
                auto_exits: list, macro: dict) -> str:
    window = "🌅 OPEN 9:35 AM ET" if datetime.utcnow().hour < 19 else "🌆 PRE-CLOSE 3:30 PM ET"
    re     = {"RISK-ON":"🟢","RISK-OFF":"🔴","CAUTION":"🟡"}.get(decision.get("regime"),"⚪️")
    ind    = macro.get("indicators", {})
    vix    = ind.get("VIX",{}).get("value")
    curve  = ind.get("T10Y2Y",{}).get("value")
    hy     = ind.get("HY_SPREAD",{}).get("value")

    lines = [
        f"📊 <b>MeritQuant Session — {window}</b>",
        f"\n{re} <b>Regime: {decision.get('regime','NEUTRAL')}</b>",
        f"💬 {decision.get('market_assessment','—')}",
    ]
    if all(x is not None for x in [vix, curve, hy]):
        lines.append(f"\n📈 VIX <b>{vix:.1f}</b>  |  Curve <b>{curve:+.2f}%</b>  |  HY <b>{hy:.0f}bps</b>")

    lines.append(f"\n🔢 <b>Actions: {len(actions)}</b>")
    for a in actions:
        e = "🟢" if a["action"]=="ENTER" else "🔴"
        lines.append(f"  {e} {a['action']} {a['ticker']} — ${a.get('position_size_usd',0):,.0f} | {a.get('conviction',0)}/10")

    if auto_exits:
        lines.append(f"\n⚡️ <b>Auto exits: {len(auto_exits)}</b>")
        for t, r, p in auto_exits:
            lines.append(f"  {'✅' if p>0 else '🛑'} {t}: {p*100:+.1f}%")

    lines += [
        f"\n💼 <b>${portfolio.get('total_value',0):,.0f}</b> total",
        f"💵 Cash: ${portfolio.get('cash',0):,.0f}",
        f"📂 Positions: {len(portfolio.get('positions',[]))}/{MAX_OPEN_POSITIONS}",
    ]
    if decision.get("memory_applied"):
        lines.append(f"\n🧠 {decision['memory_applied']}")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# 11. MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("MeritQuant AI Trader — session start")
    log.info(f"TRADE_MODE={TRADE_MODE} | UTC={datetime.utcnow().strftime('%H:%M')}")
    log.info("=" * 60)

    if not is_trade_window():
        return

    Path("data").mkdir(exist_ok=True)
    Path(REPORTS_DIR).mkdir(exist_ok=True)

    # Load state
    signals   = load_json(SIGNALS_FILE,   [])
    portfolio = load_positions()
    trade_log = load_trade_log()
    memory    = load_memory()

    if not signals:
        log.warning("No signals file — running with empty signal list.")

    # Build macro context
    macro = build_macro_context()

    # Enrich signals with balance sheet data (3-tier: EDGAR → yfinance BS → yfinance info)
    log.info("Fetching balance sheet data...")
    _load_cik_map()  # pre-warm CIK map with a single HTTP request
    for s in signals:
        bs = fetch_balance_sheet(s.get("ticker", ""))
        s["bs_score"]   = bs["score"]
        s["bs_cash"]    = round(bs["cash"] / 1e9, 2) if bs.get("cash") else None
        s["bs_de"]      = bs.get("de_ratio")
        s["bs_re"]      = bs.get("retained", [])
        s["bs_summary"] = bs.get("summary", "")
        time.sleep(0.15)  # respect SEC EDGAR rate limits

    # Refresh prices on open positions
    portfolio = update_prices(portfolio)

    # Auto SL / TP check
    portfolio, trade_log, memory, auto_exits = run_sl_tp(portfolio, trade_log, memory)
    for ticker, reason, pnl in auto_exits:
        send_telegram(sl_tp_msg(ticker, reason, pnl))
        time.sleep(1)

    # Claude Opus decision
    prompt   = build_prompt(signals, macro, portfolio, memory)
    decision = call_claude(prompt)

    if not decision:
        send_telegram("⚠️ <b>MeritQuant</b>: Claude returned no valid decision. Manual review required.")
        log.error("No valid decision — aborting.")
        return

    # Execute
    portfolio, trade_log, actions_taken = execute(decision, portfolio, trade_log)

    # Persist state
    save_json(POSITIONS_FILE, portfolio)
    sync_portfolio_json(portfolio)
    save_json(TRADE_LOG_FILE, trade_log)
    save_memory(memory)

    # Individual trade alerts
    for a in actions_taken:
        if a.get("action") in ("ENTER", "EXIT"):
            send_telegram(trade_msg(a, portfolio))
            email_alerts.send_trade_alert(a, portfolio, macro)
            time.sleep(1)

    # Backfill balance_sheet_read for positions that pre-date this field
    for pos in portfolio.get("positions", []):
        if not pos.get("balance_sheet_read"):
            bs = fetch_balance_sheet(pos.get("ticker", ""))
            pos["balance_sheet_read"] = bs.get("summary", "")
            time.sleep(0.15)

    # Generate and send reports
    ts         = datetime.utcnow().strftime("%Y%m%d_%H%M")
    pdf_bytes  = build_pdf(decision, portfolio, macro, actions_taken, auto_exits)
    xlsx_bytes = build_excel(decision, portfolio, trade_log)
    pdf_name   = f"MeritQuant_{ts}.pdf"
    xlsx_name  = f"MeritQuant_{ts}.xlsx"

    (Path(REPORTS_DIR) / pdf_name).write_bytes(pdf_bytes)
    (Path(REPORTS_DIR) / xlsx_name).write_bytes(xlsx_bytes)

    send_telegram_doc(pdf_bytes,  pdf_name,  f"📄 Trade Report {ts}")
    time.sleep(2)
    send_telegram_doc(xlsx_bytes, xlsx_name, f"📊 Trade Log {ts}")
    time.sleep(1)

    # Session summary
    send_telegram(session_msg(decision, portfolio, actions_taken, auto_exits, macro))
    email_alerts.send_session_summary(decision, portfolio, macro, actions_taken, auto_exits, pdf_bytes, xlsx_bytes)

    log.info(f"Session complete — actions={len(actions_taken)}, auto_exits={len(auto_exits)}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
